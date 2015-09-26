__author__ = 'hai'

import argparse
import os
import zipfile
import traceback
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description='Process some integers.')
parser.add_argument('--path',default='C:\\Users\\hai\\Downloads\\Hw1\\Hw1',help='path to directory to be unzip')
args = parser.parse_args()

class unzip:
    def __init__(self, path, student_file):
        self.path = path
        self.student_file = student_file
        #self.unzip()
        self.view_txt()

    def unzip(self):
        total_success = 0
        total_fail = []
        for i in os.listdir(self.path):
            tmp_fold = os.path.join(self.path,i)
            if os.path.isdir(tmp_fold):
                for j in os.listdir(tmp_fold):
                    tmp_file = os.path.join(tmp_fold, j)
                    if os.path.isfile(tmp_file) and tmp_file.endswith('.zip'):
                        print("found zip file {0}".format(tmp_file))
                        try:
                            with zipfile.ZipFile(tmp_file, 'r') as z:
                                z.extractall(tmp_fold)
                                total_success = total_success + 1
                        except:
                            print("Failed to unzip {0}".format(tmp_file))
                            total_fail.append(tmp_file)


        print("\nSummary:\n")
        print("Total zip files found: {0}".format(total_success + len(total_fail)))
        print("Success: {0}\n".format(total_success))
        print("Failed: {0}\n".format(len(total_fail)))
        for i in total_fail:
            print("{0}\n".format(i))

    def view_txt(self):
        for i in os.listdir(self.path):
            tmp_fold = os.path.join(self.path,i)
            if os.path.isdir(tmp_fold):
                 for j in os.listdir(tmp_fold):
                      tmp_file = os.path.join(tmp_fold, j)
                      if os.path.isfile(tmp_file) and tmp_file.endswith(".txt") and "output" in tmp_file.lower():
                          try:
                            print("Reading file from user {0}".format(i))
                            file = open(tmp_file, "r")
                            #print(file.read())
                            #print('\n')
                            self.grade(i)
                          except:
                              print("Error while reading file from user {0}".format(j))
                              print(traceback.format_exc())


    def grade(self, id):
        wb = load_workbook(filename = 'CS157Grade.xlsx')
        first_sheet = wb.get_sheet_names()[0]
        worksheet = wb.get_sheet_by_name(first_sheet)
        new_column = worksheet.max_row + 1
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.column == 3:
                    print(cell.value)
                    if cell.value in id:
                        print("found {0} in row {1}".format(id, cell.row))
                        #grade = input("Enter your grade for student {0}".format(id))
                        #worksheet.cell(row = cell.row, column = new_column).value=-10000
        wb.save("testexcel.xlsx")




if __name__ == "__main__":
    unzip(args.path, None)
